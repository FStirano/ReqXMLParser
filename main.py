import zipfile
import xml.etree.ElementTree as ET
from openpyxl import load_workbook
from docx import Document
from docx.oxml.table import CT_Tbl
from docx.oxml.text.paragraph import CT_P
from docx.table import Table
from docx.text.paragraph import Paragraph
from lxml import etree


# WordprocessingML namespace
NS = {
    "w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main"
}
sys_reqs = ""

class SysReq:
    def __init__(self, req_id, req_desc):
        self.req_id = req_id
        self.req_desc = req_desc
        self.req_cover = []

    def __repr__(self):
        return f"Requirement(id={self.req_id!r}, desc={self.req_desc}, data={self.req_cover})\n"
class Requirement:
    def __init__(self, module=None, iden=None, status=None, cat=None, safety=None,
                 ver_met=None, val_met=None, op=None, desc=None, cover=None):
        self.module = module
        self.iden = iden
        self.status = status
        self.cat = cat
        self.safety = safety
        self.ver_met = ver_met
        self.val_met = val_met
        self.op = op
        self.desc = desc
        self.cover = cover

    def set_module(self, value):
        self.module = value

    def get_module(self):
        return self.module

    def set_iden(self, value):
        self.iden = value

    def get_iden(self):
        return self.iden

    def set_status(self, value):
        self.status = value

    def get_status(self):
        return self.status

    def set_cat(self, value):
        self.cat = value

    def get_cat(self):
        return self.cat

    def set_safety(self, value):
        self.safety = value

    def get_safety(self):
        return self.safety

    def set_ver_met(self, value):
        self.ver_met = value

    def get_ver_met(self):
        return self.ver_met

    def set_val_met(self, value):
        self.val_met = value

    def get_val_met(self):
        return self.val_met

    def set_op(self, value):
        self.op = value

    def get_op(self):
        return self.op

    def set_desc(self, value):
        self.desc = value

    def get_desc(self):
        return self.desc

    def append_decs(self, text):
        if self.desc:
            self.desc += "\n"
        self.desc += text

    def set_cover(self, value):
        self.cover = value

    def get_cover(self):
        return self.cover

    def print(self):
        print("Requirement details:")
        print(f"\tModule: {self.module}" or f"\tModule: No module set")
        print(f"\tIdentifier: {self.iden}" or f"\tidentifier: No identifier set")
        print(f"\tStatus: {self.status}" or f"\tStatus: No status set")
        print(f"\tCategory: {self.cat}" or f"\tCategory: No category set")
        print(f"\tSafety Level (DAL): {self.safety}" or f"\tSafety Level (DAL): No safety level set")
        print(f"\tVerification Method: {self.ver_met}" or f"\tVerification Method: No method set")
        print(f"\tValidation Method: {self.val_met}" or f"\tValidation Method: No method set")
        print(f"\tOperational Mode: {self.op}" or f"\tOperational Mode: No mode set")
        print(f"\tDescription: {self.desc}" or f"\tDescription: No description set")
        print(f"\tCoverage: {self.cover}" or f"\tCoverage: No coverage set")

def load_reqs_from_excel(filename, sheetname):
    reqs = []
    sys_req_spec_file = load_workbook(filename)
    sys_req_spec_sheet = sys_req_spec_file[sheetname]
    for row in sys_req_spec_sheet.iter_rows(min_row=4, values_only=True):           # Skip header and empty rows
        req_id = row[0]         # Value in column A
        req_desc = row[1]       # Value in column B

        if not req_id:
            continue

        if isinstance(req_desc, str):
            req_desc = req_desc.replace("\n", "")

        req = SysReq(req_id, req_desc)
        reqs.append(req)

    return reqs
def get_unique_xml_tags_from_docx(docm_path):
    unique_tags = set()

    with zipfile.ZipFile(docm_path, "r") as z:
        try:
            with z.open("word/document.xml") as f:
                tree = ET.parse(f)
                root = tree.getroot()

                for elem in root.iter():
                    if '}' in elem.tag:
                        tag = elem.tag.split('}', 1)[1]
                    else:
                        tag = elem.tag
                unique_tags.add(tag)

        except ET.ParseError:
            pass

    return unique_tags
def print_tree(elem, indent=""):
    if '}' in elem.tag:
        tag = elem.tag.split('}', 1)[1]
    else:
        tag = elem.tag
    print(indent + "└── " + tag)

    for child in list(elem):
        print_tree(child, indent + "    ")
def visualize_docm_xml_tree(docm_path):
    with zipfile.ZipFile(docm_path, 'r') as z:
        try:
            with z.open("word/document.xml") as f:
                tree = ET.parse(f)
                root = tree.getroot()
                print_tree(root)
        except ET.ParseError:
            print(f"XML Parse Error")
        except Exception as e:
            print(f"Error Reading File: {e}")
def extract_text_by_paragraph_style(docm_file, target_style):
    doc = Document(docm_file)
    results = []

    for p in doc.paragraphs:
        if p.style and p.style.name == target_style:
            if p.text.strip():
                results.append(p.text.strip())

    return results
def list_styles_in_docm(docm_path):
    with zipfile.ZipFile(docm_path, "r") as z:
        styles_xml = z.read("word/styles.xml")
    root = ET.fromstring(styles_xml)

    print("Available style IDs:")
    for s in root.findall(".//w:style", NS):
        sid = s.get("{%s}styleId" % NS["w"])
        print("-", sid)
def extract_text_from_docm_by_style(docm_path, target_styles):
    results = []
    requirements = []
    currentReq = None

    with zipfile.ZipFile(docm_path, "r") as docx_zip:
        # Read the main document XML
        xml_content = docx_zip.read("word/document.xml")
        root = ET.fromstring(xml_content)

        # Iterate over all paragraphs <w:p>
        for p in root.findall(".//w:p", NS):
            # --- Check paragraph style ---
            p_style_el = p.find(".//w:pPr/w:pStyle", NS)
            paragraph_style = p_style_el.get("{%s}val" % NS["w"]) if p_style_el is not None else None

            # If paragraph style matches, extract all its text
            if paragraph_style in target_styles:
                text = "".join(t.text for t in p.findall(".//w:t", NS) if t.text)
                if text.strip():
                    results.append((paragraph_style, text))

                match paragraph_style:
                    case 'ReqTag':
                        tag = [p.strip() for p in text.split("\uf0b7")]
                        currentReq = Requirement(
                            module=tag[0].strip("[]").split("-")[1].strip(" "),
                            iden=tag[0].strip("[]"),
                            status=tag[1],
                            cat=tag[2],
                            safety=tag[3],
                            ver_met=tag[4],
                            val_met=tag[5],
                            op=tag[6]
                        )
                    case 'ReqText':
                        if currentReq:
                            if currentReq.get_desc():
                                currentReq.append_decs(text)
                            else:
                                currentReq.set_desc(text)
                    case 'ReqCover':
                        if currentReq:
                            currentReq.set_cover(text.strip("[]").split(" ")[1])
                        requirements.append(currentReq)

    return results, requirements
def check_coverage(reqs=None):
    if reqs and sys_reqs:
        for r in reqs:
            for sr in sys_reqs:
                if sr.req_id == r.cover:
                    sr.req_cover.append(r.iden)
# def iter_block_items(parent):
#    for child in parent.element.body.iterchildren():
#        if isinstance(child, CT_P):
#            yield Paragraph(child, parent)
#        elif isinstance(child, CT_Tbl):
#            yield Table(child, parent)
def extract_tables_from_docm_xml(docm_path, target_headings=None, partial=False):
    """
    Extract tables directly from the XML of a .docm file and associate
    each table with the last heading that appears above it.

    Parameters
    ----------
    docm_path : str
        Path to the .docm file.
    target_headers : list or str or None
        Restrict to tables whose heading matches these strings.
    partial : bool
        If True, header match is substring-based.

    Returns
    -------
    List of dicts:
        [
          { "header": "<text>", "table": [[...], [...]] },
          ...
        ]
    """
    if isinstance(target_headings, str):
        target_headings = [target_headings]

    # --- 1) Read the XML from the .docm zip container ---
    with zipfile.ZipFile(docm_path) as z:
        xml_content = z.read("word/document.xml")

    # --- 2) Parse XML ---
    root = etree.fromstring(xml_content)
    ns = {
        "w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main"
    }

    results = []
    current_header = None

    # --- 3) Iterate through child elements in order ---
    for child in root.xpath(".//w:body/*", namespaces=ns):

        # ---------------------------
        # A) Detect heading paragraphs
        # ---------------------------
        if child.tag == "{http://schemas.openxmlformats.org/wordprocessingml/2006/main}p":

            # Paragraph text (concatenate all runs)
            texts = child.xpath(".//w:t/text()", namespaces=ns)
            para_text = "".join(texts).strip()

            # Style (Heading 1, Heading 2, etc.)
            pstyle = child.xpath(".//w:pStyle/@w:val", namespaces=ns)
            style_name = pstyle[0] if pstyle else ""

            if style_name.startswith("Titolo") and para_text:
                current_header = para_text

        # ---------------------------
        # B) Detect tables
        # ---------------------------
        elif child.tag == "{http://schemas.openxmlformats.org/wordprocessingml/2006/main}tbl" and current_header:

            table_data = []
            for row in child.xpath(".//w:tr", namespaces=ns):
                cells = []
                for cell in row.xpath("./w:tc", namespaces=ns):
                    cell_text = "".join(cell.xpath(".//w:t/text()", namespaces=ns)).strip("[]")
                    cells.append(cell_text)
                table_data.append(cells)

            if not table_data:
                continue

            first_row = " ".join(table_data[0]).lower()

            if partial:
                row_ok = any(f.lower() in first_row for f in target_headings)
            else:
                # exact cell match: any cell equals filter text
                row_ok = any(f.lower() == cell.lower() for f in target_headings for cell in table_data[0])

            if not row_ok:
                continue

            # Store result
            results.append({
                "header": current_header.strip(" Requirements traceability"),
                "table": table_data[1:]
            })

    return results

if __name__ == '__main__':
    # Parsing Excel file to extract data
    sys_reqs = load_reqs_from_excel("System_requirements_specification__v6.xlsx", "Req.ID LED - LINKS")
    # print(reqs)

    docm_file = "FFRS_v5.docm"
    # visualize_docm_xml_tree(docm_file)

    # list_styles_in_docm(docm_file)

    styles_to_extract = {
        "ReqTag",
        "ReqText",
        "ReqCover"
    }                               # <-- use the internal style ID, not display name
                                    # Example: "Heading1", "Normal", "Quote"

    texts, reqs = extract_text_from_docm_by_style(docm_file, styles_to_extract)

    # print(f"Text using style '{styles_to_extract}':")
    # for t in texts:
    #     print("-", t)

    for r in reqs:
        r.print()

    check_coverage(reqs=reqs)
    print(sys_reqs)

    FAD_file = "FAD_v4_comments.docm"
    # visualize_docm_xml_tree(FAD_file)
    # list_styles_in_docm(FAD_file)

    tables = extract_tables_from_docm_xml(FAD_file, target_headings="Requirement Covered")
    print(tables)

#    tags = get_unique_xml_tags_from_docx(docm_file)
#    print ("Unique XML tags found in the .docm file:")
#    for t in sorted(tags):
#        print(t)



